# -*- coding: utf-8 -*-
"""Fill the cadastral "დანართი" (attachment) Excel sheet.

Reads the bundled template (sheet layout preserved: borders, dropdowns,
print area) and writes the dynamic values into the mapped cells. Values come
from the project layers + the GUI form.
"""

import os
import shutil

from qgis.core import QgsProject, QgsSettings

from . import config

TEMPLATE = os.path.join(config.ASSETS_DIR, "excel", "attachment_template.xlsx")


def openpyxl_available():
    try:
        import openpyxl  # noqa: F401
        return True
    except ImportError:
        return False


def _parse_int_person(value):
    """'givi afciauri (p/n 16001013738)' -> ('givi afciauri', '16001013738')."""
    if not value:
        return "", ""
    import re
    m = re.search(r"(.*?)\s*\(?\s*(?:p/n|p\\n|პ/ნ)?\s*(\d{6,})\)?\s*$", str(value))
    if m:
        return m.group(1).strip(" ()"), m.group(2)
    return str(value).strip(), ""


def gather_from_project(project=None):
    """Collect attachment values from the project layers + settings."""
    from .styles import normalise
    from .crs import polygon_area_m2
    project = project or QgsProject.instance()
    layers = {normalise(l.name()): l for l in project.mapLayers().values()
              if hasattr(l, "getFeatures")}

    data = {"buildings": [], "borders": [], "attendees": []}

    nak = layers.get("nakveti")
    if nak is not None:
        feat = next(nak.getFeatures(), None)
        if feat is not None:
            data["address"] = feat["ADDRESS"] if "ADDRESS" in nak.fields().names() else ""
            area = feat["Shape_Area"] if "Shape_Area" in nak.fields().names() else None
            if not area and feat.hasGeometry():
                area = polygon_area_m2(feat.geometry(), nak.crs())
            data["area"] = round(area) if area else ""
            data["designation"] = feat["FUNCTION"] if "FUNCTION" in nak.fields().names() else ""
            name, pid = _parse_int_person(
                feat["INT_PER_ID"] if "INT_PER_ID" in nak.fields().names() else "")
            data["int_person"] = name
            data["int_person_id"] = pid
            code = feat["LEGAL_DOC"] if "LEGAL_DOC" in nak.fields().names() else ""
            data["borders"] = [("A --------- B", code), ("B --------- C", code)]

    serv = layers.get("servituti")
    if serv is not None:
        f = next(serv.getFeatures(), None)
        if f is not None:
            names = serv.fields().names()
            data["obl_type"] = f["OBL_TYPE"] if "OBL_TYPE" in names else ""
            data["obl_desc"] = f["OBL_DESC"] if "OBL_DESC" in names else ""
            data["obl_area"] = round(f["Shape_Area"]) if "Shape_Area" in names and f["Shape_Area"] else ""

    bld = layers.get("shenoba")
    if bld is not None:
        names = bld.fields().names()
        for f in bld.getFeatures():
            data["buildings"].append({
                "num": f["BLD_NUM"] if "BLD_NUM" in names else "",
                "func": f["FUNCTION"] if "FUNCTION" in names else "",
                "state": f["STATE"] if "STATE" in names else "",
                "floors": f["FLOORS"] if "FLOORS" in names else "",
                "area": round(f["Shape_Area"]) if "Shape_Area" in names and f["Shape_Area"] else "",
            })

    s = QgsSettings()
    g = config.SETTINGS_GROUP
    data["auth_legal"] = s.value(f"{g}/auth_legal", "")
    data["auth_id"] = s.value(f"{g}/auth_id", "")
    data["auth_contact"] = s.value(f"{g}/auth_contact", "")
    data["auth_person"] = s.value(f"{g}/auth_person", "")
    return data


def write_attachment(out_path, data, template=None):
    """Write filled attachment to out_path. Raises ImportError if openpyxl
    is missing (the caller shows an install hint)."""
    import openpyxl
    template = template or TEMPLATE
    if os.path.abspath(out_path) != os.path.abspath(template):
        shutil.copyfile(template, out_path)
    wb = openpyxl.load_workbook(out_path)
    ws = wb[config.EXCEL_SHEET] if config.EXCEL_SHEET in wb.sheetnames else wb.worksheets[0]
    C = config.EXCEL_CELLS

    def put(cell, value):
        if value not in (None, ""):
            ws[cell] = value

    put(C["address"], data.get("address"))
    put(C["area"], data.get("area"))
    put(C["designation"], data.get("designation"))

    # buildings
    r = C["building_first_row"]
    cols = C["building_cols"]
    for b in data.get("buildings", []):
        if r > C["building_last_row"]:
            break
        put(f"{cols['num']}{r}", b.get("num"))
        put(f"{cols['func']}{r}", b.get("func"))
        put(f"{cols['state']}{r}", b.get("state"))
        put(f"{cols['floors']}{r}", b.get("floors"))
        put(f"{cols['area']}{r}", b.get("area"))
        r += 1

    put(C["obl_desc"], data.get("obl_desc"))
    put(C["obl_type"], data.get("obl_type"))
    put(C["obl_area"], data.get("obl_area"))

    put(C["lin_type"], data.get("lin_type"))
    put(C["lin_actual"], data.get("lin_actual"))
    put(C["lin_planned"], data.get("lin_planned"))
    put(C["lin_pt_type"], data.get("lin_pt_type"))
    put(C["lin_pt_count"], data.get("lin_pt_count"))

    # borders
    r = C["border_first_row"]
    bc = C["border_cols"]
    for seg, neigh in data.get("borders", []):
        put(f"{bc['segment']}{r}", seg)
        put(f"{bc['neighbour']}{r}", neigh)
        r += 1
    # attendees
    r = C["border_first_row"]
    ac = C["attendee_cols"]
    for att in data.get("attendees", []):
        put(f"{ac['name']}{r}", att.get("name"))
        put(f"{ac['signature']}{r}", att.get("signature"))
        r += 1

    put(C["auth_legal"], data.get("auth_legal"))
    put(C["auth_id"], data.get("auth_id"))
    put(C["auth_contact"], data.get("auth_contact"))
    put(C["method"], data.get("method"))
    put(C["survey_date"], data.get("survey_date"))
    put(C["auth_person"], data.get("auth_person"))
    put(C["int_person"], data.get("int_person"))
    put(C["int_person_id"], data.get("int_person_id"))

    wb.save(out_path)
    return out_path
